import barcode
from barcode.writer import ImageWriter
import openpyxl
import os

#Incluir la localizacion del fichero excel con los datos de Zooplus:
directory = os.getcwd()
path = directory + "\\Precios.xlsx"
#print(directory)

# Se abre el workbook y la hoja seleccionados (donde se guardo el documento por ultima vez)
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Obtener el numero de filas y columnas usadas
row_limit = sheet_obj.max_row # type: ignore
column_limit = sheet_obj.max_column # type: ignore
  
print("Total de filas:", row_limit)
print("Total de columnas:", column_limit)

#Introducir el numero de columna en la que se encuentra el EAN (A=1, B=2, C=3...)
ean_col = 1

#Introducir el numero de la fila en la que se encuentra el primer valor de EAN
first_ean_row = 2

for i in range(first_ean_row, row_limit + 1): 
    cell_obj = sheet_obj.cell(row = i, column = ean_col) # type: ignore
    ean = str(cell_obj.value)

# Definir los datos con el formato de GS1-128 (AIs + datos)
# Aquí, 01 es el AI para GTIN, seguido de un GTIN ficticio "01234567891234"
    gs1_data = ean

# Genera el código de barras en formato Code128 (base para GS1-128)
    code128 = barcode.get('code128', gs1_data, writer=ImageWriter())

# Guarda la imagen
    code128.save(f"{directory}\\codigo_gs1_128_n{i}")