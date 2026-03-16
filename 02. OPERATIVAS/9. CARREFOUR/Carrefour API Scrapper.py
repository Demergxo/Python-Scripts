import requests
import urllib3
import openpyxl
import time
from alive_progress import alive_bar
from alive_progress import animations



urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def quit_two_firsts(string):
    return string[2:]

#Incluir la localizacion del fichero excel con los datos de carrefour:
directory = r"C:\Users\jgmeras\Documents\Python Scripts\OPERATIVAS\9. CARREFOUR" #os.getcwd()
path = directory + r"\Precios.xlsx"
#print(directory)

# Se abre el workbook y la hoja seleccionados (donde se guardo el documento por ultima vez)
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Obtener el numero de filas y columnas usadas
row_limit = sheet_obj.max_row # type: ignore
column_limit = sheet_obj.max_column # type: ignore
  
print("Total de filas:", row_limit)
print("Total de columnas:", column_limit)

#Introducir el numero de columna en la que se encuentra el EAN (A=1, B=2, C=3...)
ean_col = 1

#Introducir el numero de la fila en la que se encuentra el primer valor de EAN
first_ean_row = 2

#Introducir el numero de columna en la que se encuentra el precio
precio_col = 3


#Configurar navegador

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)"
                  " Chrome/116.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Referer": "https://www.carrefour.es/",
    "Origin": "https://www.carrefour.es"
}

url = "https://www.carrefour.es/search-api/query/v1/search"

#Bucle para leer códigos del Excel

for i in range(first_ean_row, row_limit + 1): 
    cell_obj = sheet_obj.cell(row = i, column = ean_col) # type: ignore
    ean1 = str(cell_obj.value)
    codigo = quit_two_firsts(ean1)
    
    
    if codigo is None:
        codigo = ""

#Configuración de la consulta a la API        
    params = {
        "internal": "true",
        "query": codigo,
        "origin": "url:external",
        "start": "0",
        "rows": "24",
        "instance": "x-carrefour",
        "env": "https://www.carrefour.es",
        "scope": "desktop",
        "lang": "es",
        "session": "empathy",
        "citrusCatalog": "home",
        "baseUrlCitrus": "https://www.carrefour.es",
        "enabled": "true",
        "store": "005290",
        "shopperId": "30VTw0c7YwPy8SJKjSuu9xKbCE1",
        "hasConsent": "true",
        "siteKey": "wFOzqveg",
        "grid_def_search_sponsor_product": "3,5,11,13,19",
        "grid_def_search_butterfly_banner": "7-8,15-16",
        "grid_def_search_sponsor_product_tablet": "2,4,11,13,19",
        "grid_def_search_butterfly_banner_tablet": "6,12",
        "grid_def_search_sponsor_product_mobile": "2,4,11,13,19",
        "grid_def_search_butterfly_banner_mobile": "6,12",
        "grid_def_search_luckycart_banner": "22"
    }
    
    #Realización de la consulta

    print(f"\n🔍{i}. Consultando código: {codigo}")
    response = requests.get(url, params=params, headers=headers, verify=False)


    try:
        if response.status_code == 200:
            try:
                data = response.json()
                docs = data.get("content", {}).get("docs", [])
                if docs:
                    precio = docs[0].get("active_price", "Precio no disponible")
                    #nombre = docs[0].get("name", "Nombre no disponible")
                    print(f"💶 Precio: {precio} €")
                    c1 = sheet_obj.cell(row = i, column = precio_col) # type: ignore
                    c1.value = precio
                else:
                    print("⚠️ Producto no encontrado.")
            except Exception as e:
                print(f"❌ Error al procesar JSON: {e}")
        else:
            print(f"❌ Error HTTP {response.status_code}")
    except ConnectionError:
        #Manejo de caida de conexión
        bar = animations.bar_factory('😴', tip="😪", background='zZz', borders=('Durmiendo 👉 ->|','|<- Terminado 🤘'), errors=('<---👀', '💀'))
        print("Hemos tenido un error de conexión, voy a descansar 10 segundos. ZZZZZZzzzzz...")
        total = 10
        i = 0
        
        with alive_bar(total, tittle = "Durmiendo", bar=bar) as bar:
            while i < total:
                time.sleep(1)
                bar()
                i += 1
        continue
        
        
wb_obj.save(path)
print("ARCHIVO GUARDADO / PROGRAMA FINALIZADO")



