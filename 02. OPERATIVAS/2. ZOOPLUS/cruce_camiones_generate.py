import openpyxl

def dividir_en_camiones(archivo_origen, archivo_destino):
    # Cargar el archivo de origen
    origen_wb = openpyxl.load_workbook(archivo_origen, data_only=True)
    origen_sheet = origen_wb.active

    columnas_necesarias = ["Ubicación", "LA", "Descripción", "Soporte"]

    # Obtener los encabezados
    encabezados_origen = {origen_sheet.cell(row=1, column=col).value: col for col in range(1, origen_sheet.max_column + 1)}

    # Verificar si todas las columnas necesarias están presentes
    if not all(col in encabezados_origen for col in columnas_necesarias):
        print("❌ Faltan columnas necesarias en el archivo de origen.")
        return

    # Obtener los índices de las columnas necesarias
    columnas_origen = [encabezados_origen[col] for col in columnas_necesarias]

    # Extraer los datos a partir de la fila 2
    data = [tuple(row[col - 1] for col in columnas_origen) for row in origen_sheet.iter_rows(min_row=2, max_row=origen_sheet.max_row, values_only=True) if any(row)]

    if not data:
        print("❌ No hay datos para procesar.")
        return

    print(f"✅ Datos extraídos: {len(data)} filas")

    # Crear el nuevo archivo de destino
    destino_wb = openpyxl.Workbook()
    destino_wb.remove(destino_wb.active)  # Eliminar la hoja predeterminada

    # Dividir los datos en grupos de 33 filas por hoja
    for i, inicio in enumerate(range(0, len(data), 33), start=1):
        nombre_pestaña = f"Camión {i:02}"
        destino_sheet = destino_wb.create_sheet(title=nombre_pestaña)

        # Escribir los encabezados en la fila 1
        for col, encabezado in enumerate(columnas_necesarias, start=1):
            destino_sheet.cell(row=1, column=col, value=encabezado)

        # Escribir los datos a partir de la fila 2
        for fila, row_data in enumerate(data[inicio:inicio + 33], start=2):
            for col, value in enumerate(row_data, start=1):
                destino_sheet.cell(row=fila, column=col, value=value)

        print(f"📄 Creada pestaña: {nombre_pestaña} con {len(data[inicio:inicio + 33])} filas")

    # Guardar el archivo de destino
    try:
        destino_wb.save(archivo_destino)
        print(f"✅ Archivo generado exitosamente: {archivo_destino}")
    except Exception as e:
        print(f"❌ Error al guardar el archivo: {e}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Uso: python dividir_camiones.py <archivo_origen.xlsx> <archivo_destino.xlsx>")
    else:
        dividir_en_camiones(sys.argv[1], sys.argv[2])
