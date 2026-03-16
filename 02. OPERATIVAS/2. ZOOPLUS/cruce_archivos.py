import openpyxl

def copiar_datos(archivo_origen, archivo_destino):
    # Cargar los archivos
    origen_wb = openpyxl.load_workbook(archivo_origen, data_only=True)
    destino_wb = openpyxl.load_workbook(archivo_destino)

    columnas_necesarias = ["Ubicación", "LA", "Descripción", "Soporte"]

    for i in range(1, 61):
        origen_sheet_name = f"Camión {i:02}"
        destino_sheet_name = f"Camion {i:02}"

        if origen_sheet_name in origen_wb.sheetnames and destino_sheet_name in destino_wb.sheetnames:
            origen_sheet = origen_wb[origen_sheet_name]
            destino_sheet = destino_wb[destino_sheet_name]

            # Obtener encabezados
            encabezados_origen = {origen_sheet.cell(row=1, column=col).value: col for col in range(1, origen_sheet.max_column + 1)}
            encabezados_destino = {destino_sheet.cell(row=7, column=col).value: col for col in range(1, destino_sheet.max_column + 1)}

            if not all(col in encabezados_origen for col in columnas_necesarias) or not all(col in encabezados_destino for col in columnas_necesarias):
                print(f"⚠️ Columnas faltantes en {origen_sheet_name} o {destino_sheet_name}, saltando...")
                continue

            # Índices de columnas
            columnas_origen = [encabezados_origen[col] for col in columnas_necesarias]
            columnas_destino = [encabezados_destino[col] for col in columnas_necesarias]

            # Extraer los datos
            data = [tuple(row[col - 1] for col in columnas_origen) for row in origen_sheet.iter_rows(min_row=2, max_row=origen_sheet.max_row, values_only=True) if any(row)]

            if not data:
                print(f"⚠️ No hay datos en {origen_sheet_name}, saltando...")
                continue

            # Encontrar la primera fila vacía
            first_empty_row = 8
            while destino_sheet.cell(row=first_empty_row, column=1).value is not None:
                first_empty_row += 1

            # Escribir los datos, respetando el límite de 33 líneas por hoja
            for row_data in data[:33]:
                for j, value in enumerate(row_data):
                    destino_sheet.cell(row=first_empty_row, column=columnas_destino[j], value=value)
                first_empty_row += 1

    # Guardar el archivo como xlsx
    try:
        destino_wb.save(archivo_destino)
        print(f"✅ Datos copiados exitosamente a {archivo_destino}")
    except Exception as e:
        print(f"❌ Error al guardar el archivo: {e}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Uso: python cruce_camiones.py <archivo_origen.xlsx> <archivo_destino.xlsx>")
    else:
        copiar_datos(sys.argv[1], sys.argv[2])
