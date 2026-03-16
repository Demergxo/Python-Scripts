from openpyxl import load_workbook
from openpyxl.styles import Protection
import re
import sys
import os

usuario = os.environ['USERNAME']

def proteger_excel(archivo_entrada, archivo_salida, contraseña):
    # Cargar el archivo
    wb = load_workbook(archivo_entrada)

    # Expresión regular para identificar hojas "Camion 01" a "Camion 45"
    patron_camion = re.compile(r"^Camion (0[1-9]|[1-3][0-9]|45)$")

    # Recorrer todas las hojas y proteger solo las que cumplen con "Camion xx"
    for hoja in wb.sheetnames:
        if patron_camion.match(hoja):  # Verificar si el nombre coincide
            ws = wb[hoja]

            # Bloquear todas las celdas por defecto
            for fila in ws.iter_rows():
                for celda in fila:
                    celda.protection = Protection(locked=True)

            # Desbloquear solo el rango E8:E40
            for fila in range(8, 41):  # Desde la fila 8 hasta la 40
                ws[f"E{fila}"].protection = Protection(locked=False)

            # Proteger la hoja con contraseña
            ws.protection.sheet = True
            ws.protection.password = contraseña

    # Guardar el archivo con las hojas protegidas
    wb.save(archivo_salida)
    print(f"Archivo guardado como: {archivo_salida}")
archivo_salida = f"C:\\Users\\{usuario}\\OneDrive - GXO\\Escritorio\\archivo_protegido.xlsx"
contraseña = "1111"


# Ejecutar la función
if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python bloqueado_masivo.py <archivo1.xlsx>")
    else:
        proteger_excel(sys.argv[1], archivo_salida, contraseña)
        




