import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

#pip install alive-progress
#pip install colorama
from alive_progress import alive_bar
from colorama import init, Fore, Back, Style


init()

def procesar_excel(ruta_archivo):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(filename=ruta_archivo)
    sheet = wb.active

    # Diccionario para almacenar los precios máximos por LA
    precios_maximos = {}
    cambios_descripcion = {}

    # Lista para almacenar todos los registros únicos
    registros_unicos = []

    # Iterar sobre las filas del archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True): #type: ignore
        LA, descripcion, precio, revisado = row
        
        # Verificar cambios en la descripción
        if LA in cambios_descripcion:
            while True:
                if cambios_descripcion[LA] != descripcion:
                    print("\n"+Fore.RED+f"[!] Se ha encontrado un cambio en la descripción del LA {LA}:"+Fore.WHITE)
                    print(f"\n[?] Antigua descripción: {cambios_descripcion[LA]}")
                    print(f"[?] Nueva descripción: {descripcion}")
                    respuesta = input("\n[?]¿Desea utilizar la nueva descripción? (S/N): ")
                    if respuesta.lower() == "s":
                        cambios_descripcion[LA] = descripcion
                        revisado = None
                        
                        break
                    if respuesta.lower() == 'n':
                        break
                    else:
                        print("\n"+Back.RED+"[!] Opción incorrecta, elija S o N"+Back.BLACK)
                else:
                    break
        else:
            cambios_descripcion[LA] = descripcion
        
        # Verificar si el LA ya está en el diccionario de precios máximos
        if LA in precios_maximos:
            # Si el LA ya tiene una entrada en precios_maximos, comprobar si es revisado o no
            if revisado == "X":
                precios_maximos[LA] = precio
                
            else:
                precios_maximos[LA] = max(float(precio), float(precios_maximos[LA]))#type: ignore
        else:
            precios_maximos[LA] = precio


       
        # Almacenar el registro actual si es único o tiene revisado "X"
        if LA not in [r[0] for r in registros_unicos] or revisado == "X":
            #print("LA {} desc {} rev {}".format(LA, cambios_descripcion[LA], revisado))
            registros_unicos.append((LA, descripcion, precio, revisado))
    
    return precios_maximos, registros_unicos, cambios_descripcion

def create_table(path, wb_obj):
    #Comprobamos si hay una hoja creada, si no la hay, se crea y se genera la tabla
    
    if "Resultado" in wb_obj.sheetnames:
    
        print("\n[!] Tabla Resultado creada, bórrela y vuelva a lanzar el código")
    
    else:
        #Crea la hoja
        ws1 = wb_obj.create_sheet("Resultado")
        #Genera las cabeceras
        ws1['A1'] = "LA"
        ws1['B1'] = "Descripcion"
        ws1['C1'] = "Precio"
        ws1['D1'] = "Revisado"
        #genera la tabla
        tab = Table(displayName="Resultado", ref="A1:D{}".format(row_limit))
        #Le da un estilo a la tabla
        style = TableStyleInfo(name="TableStyleMedium10", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        #añade la tabla y guarda
        ws1.add_table(tab)
        wb_obj.save(path)
        print("\n[+] Añadida un nueva Hoja 'Resultado' y una tabla con los datos")
        
def input_values_table(registro, la_col, descripcion_col, precio_col, revisado_col, wb_obj, cambios_descripcion, precios_maximos, registros):
    # Insertamos los valores en cada columna
    sheet_result = wb_obj['Resultado']
        
    next_row = 1
    while sheet_result.cell(row=next_row, column=la_col).value:
        next_row += 1
    
    c1 = sheet_result.cell(row=next_row, column=la_col)
    c1.value = str(registro[0])
    
    # Verificar si hay una nueva descripción para este LA
    if registro[0] in cambios_descripcion:
        c1 = sheet_result.cell(row=next_row, column=descripcion_col)
        c1.value = cambios_descripcion[registro[0]]
        
        
        # Actualizar la descripción en la lista de registros
        indice = next(i for i, r in enumerate(registros) if r[0] == registro[0])
        registros[indice] = (registros[indice][0], cambios_descripcion[registro[0]], registros[indice][2], registros[indice][3])
    else:
        c1 = sheet_result.cell(row=next_row, column=descripcion_col)
        c1.value = str(registro[1])
    
    # Copiar el precio máximo solo si no está revisado (sin "X" en la columna revisado)
    if registro[3] != "X":
        c1 = sheet_result.cell(row=next_row, column=precio_col)
        c1.value = precios_maximos.get(registro[0], 0)  # Obtener el precio máximo del diccionario precios_maximos
    else:
        c1 = sheet_result.cell(row=next_row, column=precio_col)
        c1.value = str(registro[2])  # Copiar el precio de la fila
    
    # Copiar la columna revisada
    c1 = sheet_result.cell(row=next_row, column=revisado_col)
    c1.value = str(registro[3])
    
    
#Incluir la localizacion del fichero excel con los datos de Zooplus:
directory = os.getcwd()
path = directory + "\\2. ZOOPLUS\\Filtrado_precios.xlsx" #\\2. ZOOPLUS
#print(directory)
with alive_bar(1) as bar:   # default setting
    for i in range(1):
        #time.sleep(0.03)
        # Se abre el workbook y la hoja seleccionados (donde se guardo el documento por ultima vez)
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        # Obtener el numero de filas y columnas usadas
        row_limit = sheet_obj.max_row #type: ignore
        column_limit = sheet_obj.max_column #type: ignore
        
        #Introducir el numero de columna en la que se encuentra el EAN (A=1, B=2, C=3...)
        la_col = 1

        #Introducir el numero de la fila en la que se encuentra el primer valor de EAN
        first_la_row = 2

        #Introducir el numero de columna en la que se encuentra el precio
        precio_col = 3
        
        #Opciones de revision
        revisado_col = 4
        
        #Descripcion
        descripcion_col = 2
        
        bar()
print("""
      
███████╗██╗██╗  ████████╗██████╗  █████╗ ██████╗  ██████╗         ██████╗ ███████╗        ██████╗ ██████╗ ███████╗ ██████╗██╗ ██████╗ ███████╗
██╔════╝██║██║  ╚══██╔══╝██╔══██╗██╔══██╗██╔══██╗██╔═══██╗        ██╔══██╗██╔════╝        ██╔══██╗██╔══██╗██╔════╝██╔════╝██║██╔═══██╗██╔════╝
█████╗  ██║██║     ██║   ██████╔╝███████║██║  ██║██║   ██║        ██║  ██║█████╗          ██████╔╝██████╔╝█████╗  ██║     ██║██║   ██║███████╗
██╔══╝  ██║██║     ██║   ██╔══██╗██╔══██║██║  ██║██║   ██║        ██║  ██║██╔══╝          ██╔═══╝ ██╔══██╗██╔══╝  ██║     ██║██║   ██║╚════██║
██║     ██║███████╗██║   ██║  ██║██║  ██║██████╔╝╚██████╔╝        ██████╔╝███████╗        ██║     ██║  ██║███████╗╚██████╗██║╚██████╔╝███████║
╚═╝     ╚═╝╚══════╝╚═╝   ╚═╝  ╚═╝╚═╝  ╚═╝╚═════╝  ╚═════╝         ╚═════╝ ╚══════╝        ╚═╝     ╚═╝  ╚═╝╚══════╝ ╚═════╝╚═╝ ╚═════╝ ╚══════╝
                                                                                                                                              
      
      """)
print("\n"+Back.BLUE+"[+] Analisis de Excel"+Back.BLACK+"\n")

print(Back.BLACK+Fore.GREEN+"[+] Total de filas:", row_limit)
print("[+] Total de columnas:", column_limit)
create_table(path, wb_obj)

print("\n"+Back.BLUE+"[+] Comienzo de analisis de precios"+Back.BLACK+"\n")
       
#Cogemos el primer LA, descripcion 
precios_maximos, registros, cambios_descripcion = procesar_excel(path)
i = 1

with alive_bar(len(registros)) as bar:   
    
    for registro in registros:
        input_values_table(registro, la_col, descripcion_col, precio_col, revisado_col, wb_obj, cambios_descripcion, precios_maximos, registros)
        i = i+1
        bar()
    
wb_obj.save(path)
print("\n"+Back.GREEN+"[+] Proceso terminado"+Back.BLACK+"\n")                           


