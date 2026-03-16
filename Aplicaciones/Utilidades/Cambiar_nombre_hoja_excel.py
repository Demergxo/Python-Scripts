import os
from openpyxl import load_workbook

# Ruta a la carpeta con los archivos .xlsx
CARPETA = r'C:\Users\jgmeras\OneDrive - GXO\Escritorio\Ajustes Altura\SbL'

# Nombre actual y nuevo de la hoja
NOMBRE_ANTIGUO = "Stockporubicación"
NUEVO_NOMBRE = "Stockporubicacin"

def renombrar_hoja(archivo):
    try:
        wb = load_workbook(archivo)
        if NOMBRE_ANTIGUO in wb.sheetnames:
            wb[NOMBRE_ANTIGUO].title = NUEVO_NOMBRE
            wb.save(archivo)
            return True  # Se hizo el cambio
        else:
            return False  # La hoja no existe
    except Exception as e:
        print(f"\n❌ Error al procesar {archivo}: {e}")
        return False

def main():
    archivos_xlsx = [f for f in os.listdir(CARPETA) if f.lower().endswith(".xlsx")]
    total_archivos = len(archivos_xlsx)
    modificados = []

    for idx, nombre_archivo in enumerate(archivos_xlsx, start=1):
        print(f"Procesando {idx}/{total_archivos}: {nombre_archivo}")
        ruta_completa = os.path.join(CARPETA, nombre_archivo)

        cambiado = renombrar_hoja(ruta_completa)
        if cambiado:
            modificados.append(nombre_archivo)

    print("\n📋 Resumen:")
    if modificados:
        print("Se renombró la hoja en los siguientes archivos:")
        for archivo in modificados:
            print(" -", archivo)
    else:
        print("✅ Ningún archivo requería cambios (la hoja no existía o ya estaba bien nombrada).")

if __name__ == "__main__":
    main()
