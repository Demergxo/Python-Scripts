import os
from openpyxl import load_workbook

# Ruta a la carpeta con los archivos .xlsx
CARPETA = r'C:\Users\jgmeras\OneDrive - GXO\Escritorio\Ajustes Altura\SbL'

# Lista de referencia con los nombres esperados en las celdas B6 a X6

columnas_referencia = [
    "Pasillo", "Cara", "Columna", "Altura", "Posición", "Geo. Zone", "N° Support",
    "Artículo", "EAN", "Designación", "Descripción larga", "Familia Almacenamiento", 
    "MegaPack", "Unids/Pallet", "Unds. Totales", "Peso Neto", "Alto", "Largo", "Fondo",
    "Condition", "Condition Desc.", "Reception Date", "Exp. date"

]

def revisar_columnas(archivo, columnas_referencia):
    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        # Extraer valores desde B6 a X6 (columna 2 a 24)
        encabezados = [ws.cell(row=6, column=col).value for col in range(2, 25)] #type: ignore
        son_iguales = encabezados == columnas_referencia
        return son_iguales, encabezados
    except Exception as e:
        print(f"\n❌ Error al procesar {archivo}: {e}")
        return False, []

def main():
    archivos_diferentes = []
    archivos_xlsx = [f for f in os.listdir(CARPETA) if f.lower().endswith(".xlsx")]
    total_archivos = len(archivos_xlsx)

    for idx, nombre_archivo in enumerate(archivos_xlsx, start=1):
        print(f"Procesando {idx}/{total_archivos}: {nombre_archivo}")
        ruta_completa = os.path.join(CARPETA, nombre_archivo)

        es_igual, encabezados = revisar_columnas(ruta_completa, columnas_referencia)
        if not es_igual:
            archivos_diferentes.append((nombre_archivo, encabezados))

    print("\nResumen:")
    if archivos_diferentes:
        print("Archivos con columnas distintas a las de referencia:")
        for archivo, columnas in archivos_diferentes:
            print(f" - {archivo}")
    else:
        print("✅ Todos los archivos tienen las columnas correctas.")

if __name__ == "__main__":
    main()