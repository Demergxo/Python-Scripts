import pandas as pd
import openpyxl
import os

indice = {
    "Dogs": [],
    "Cats": [],
    "Others": [],
    "Out of family": [],
}

def leer_master_data():
    # Incluir la localizacion del fichero excel con los datos de Zooplus:
    directory = os.getcwd()
    path = os.path.join(directory, "6. CABANILLAS", "Mturn", "Master_data.xlsx")
    
    # Se abre el workbook y la hoja seleccionados (donde se guardo el documento por ultima vez)
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    # Obtener el numero de filas y columnas usadas
    row_limit = sheet_obj.max_row
    column_limit = sheet_obj.max_column
    
    #print("Total de filas:", row_limit)
    #print("Total de columnas:", column_limit)
    
    # Comienzo datos
    min_row = 2
    
    # Definición de Familia
    family_col = 4
    
    # Columna de La's
    la_col = 1
    
    for i in range(min_row, row_limit + 1):
        cell_obj = sheet_obj.cell(row=i, column=family_col)
        family = str(cell_obj.value).strip()
        la_obj = sheet_obj.cell(row=i, column=la_col)
        la = str(la_obj.value).strip()

        if family == "Dogs":
            indice["Dogs"].append(la)
        elif family == "Cats":
            indice["Cats"].append(la)
        elif family == "Others":
            indice["Others"].append(la)
        else:
            indice["Out of family"].append(la)

    #print("Índice cargado:", indice)  # Depuración
    return indice        

def separar_por_delivery():
    # Leer el archivo Excel
    directory = os.getcwd()
    path = os.path.join(directory, "6. CABANILLAS", "Mturn", "Inbound_Follow-up.xlsx")
    sheet_name = "Inbound Follow-up"
    df = pd.read_excel(path, sheet_name=sheet_name, skiprows=3)

    df = df.drop(columns=df.columns[0])

    # Asegurarnos de que las columnas 'Delivery ID' y 'Item' existan
    if 'Delivery ID' not in df.columns or 'Item' not in df.columns:
        raise ValueError("El archivo Excel debe contener las columnas 'Delivery ID' y 'Item'.")

    # Agrupar por 'Delivery ID' y crear una lista de 'Item' (SKU) para cada 'Delivery ID'
    result = df.groupby('Delivery ID')['Item'].apply(list).reset_index()

    # Convertir el resultado en un diccionario para usarlo en otra función
    delivery_sku_dict = result.set_index('Delivery ID')['Item'].to_dict()

    #print("Delivery SKU dict:", delivery_sku_dict)  # Depuración
    return delivery_sku_dict    

def catalogar_delivery_skus(delivery, index):
    separator = {}
    
    for delivery_id, skus in delivery.items():
        dogs = 0
        cats = 0
        others = 0
        #print(f"Procesando delivery ID: {delivery_id}")  # Depuración
        for sku in skus:
            sku = str(sku).strip()  # Asegurarse de que no hay espacios en blanco
            #print(f"  SKU: {sku}")  # Depuración
            if sku in index["Dogs"]:
                dogs += 1
                #print(f"    {sku} es un Dog SKU")  # Depuración
            elif sku in index["Cats"]:
                cats += 1
                #print(f"    {sku} es un Cat SKU")  # Depuración
            elif sku in index["Others"]:
                others += 1
                #print(f"    {sku} es un Other SKU")  # Depuración
        total = dogs + cats + others
        if total > 0:
            dogs_percentage = (dogs / total) * 100
            cats_percentage = (cats / total) * 100
            others_percentage = (others / total) * 100

            if dogs >= cats and dogs >= others:
                majority_family = "Dogs"
                majority_percentage = dogs_percentage
            elif cats >= dogs and cats >= others:
                majority_family = "Cats"
                majority_percentage = cats_percentage
            else:
                majority_family = "Others"
                majority_percentage = others_percentage
            
            separator[delivery_id] = {
                "Dogs": dogs,
                "Cats": cats,
                "Others": others,
                "Majority Family": majority_family,
                "Majority Percentage": majority_percentage
            }
    
    #print("Separator:", separator)  # Depuración
    return separator

indice = leer_master_data()
delivery = separar_por_delivery()
sep = catalogar_delivery_skus(delivery, indice)

for delivery_id, details in sep.items():
    print(f"La delivery [{delivery_id}] tiene la mayor parte de [{details['Majority Family']}] con un [{details['Majority Percentage']:.2f}%]")


