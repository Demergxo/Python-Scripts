def leer_master_data(indice):
    import os
    import openpyxl
    # Incluir la localizacion del fichero excel con los datos de Zooplus:
    directory = os.getcwd()
    path = os.path.join(directory, "6. CABANILLAS", "Mturn", "Muelle_optimo", "Master_data.xlsx")
    
    # Se abre el workbook y la hoja seleccionados (donde se guardo el documento por ultima vez)
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    # Obtener el numero de filas y columnas usadas
    row_limit = sheet_obj.max_row #type: ignore
    column_limit = sheet_obj.max_column #type: ignore
    
    print("Total de filas:", row_limit)
    print("Total de columnas:", column_limit)
    
    # Comienzo datos
    min_row = 2
    
    # Definición de Familia
    family_col = 3
    
    # Columna de La's
    la_col = 1
    
    for i in range(min_row, row_limit + 1): #type: ignore
        cell_obj = sheet_obj.cell(row=i, column=family_col) #type: ignore
        family = str(cell_obj.value).strip()
        la_obj = sheet_obj.cell(row=i, column=la_col) #type: ignore
        la = str(la_obj.value).strip()

        if family == "Dogs":
            indice["Dogs"].append(la)
        elif family == "Cats":
            indice["Cats"].append(la)
        elif family == "Others":
            indice["Others"].append(la)
        elif family == "Consumables":
            indice["Consumables"].append(la)
        else:
            indice["Out of family"].append(la)

    #print("Índice cargado:", indice)  # Depuración
    return indice        

def separar_por_delivery():
    import os
    import pandas as pd
    # Leer el archivo Excel
    directory = os.getcwd()
    path = os.path.join(directory, "6. CABANILLAS", "Mturn", "Muelle_optimo", "Inbound_Follow-up.xlsx")
    sheet_name = "Inbound Follow-up"
    df = pd.read_excel(path, sheet_name=sheet_name, skiprows=3)

    df = df.drop(columns=df.columns[0])

    # Asegurarnos de que las columnas 'Delivery ID' y 'Item' existan
    if 'Delivery ID' not in df.columns or 'Item' not in df.columns or 'Quantity Expected' not in df.columns:
        raise ValueError("El archivo Excel debe contener las columnas 'Delivery ID', 'Item' y 'Quantity Expected'.")

    # Agrupar por 'Delivery ID' y crear una lista de 'Item' (SKU) para cada 'Delivery ID'
    result = df.groupby('Delivery ID').apply(lambda x: list(zip(x['Item'], x['Quantity Expected']))).reset_index()

    # Convertir el resultado en un diccionario para usarlo en otra función
    delivery_sku_dict = result.set_index('Delivery ID')[0].to_dict()

    #print("Delivery SKU dict:", delivery_sku_dict)  # Depuración
    return delivery_sku_dict 




def catalogar_delivery_skus(delivery, index):
    separator = {}
    
    for delivery_id, skus in delivery.items():
        dogs = 0
        cats = 0
        others = 0
        cons = 0
        no_fam = 0
        #print(f"Procesando delivery ID: {delivery_id}")  # Depuración
        for sku in skus:
            sku = str(sku).strip()  # Asegurarse de que no hay espacios en blanco
            #print(f"  SKU: {sku}")  # Depuración
            if sku in index["Dogs"]:
                dogs += 1
                #print(f"    {sku} es un Dog SKU")  # Depuración
            elif sku in index["Cats"]:
                cats += 1
                #print(f"    {sku} es un Cat SKU")  # Depuración
            elif sku in index["Others"]:
                others += 1
                #print(f"    {sku} es un Other SKU")  # Depuración
            elif sku in index["Consumables"]:
                cons += 1
            else:
                no_fam += 1
        total = dogs + cats + others + no_fam
        if total > 0:
            dogs_percentage = (dogs / total) * 100
            cats_percentage = (cats / total) * 100
            others_percentage = (others / total) * 100
            consumables_percentage = (cons / total) * 100

            if dogs >= cats and dogs >= others:
                majority_family = "Dogs"
                majority_percentage = dogs_percentage
            elif cats >= dogs and cats >= others:
                majority_family = "Cats"
                majority_percentage = cats_percentage
            elif cons >= dogs and cons >= cats:
                majority_family = "Consumables"
                majority_percentage = consumables_percentage
            else:
                majority_family = "Others"
                majority_percentage = others_percentage
            
            separator[delivery_id] = {
                "Dogs": dogs,
                "Cats": cats,
                "Consumables": cons,
                "Others": others,
                "Majority Family": majority_family,
                "Majority Percentage": majority_percentage
            }
    
    #print("Separator:", separator)  # Depuración
    return separator

def leer_archivo_implantaciones():
    import pandas as pd
    import os
    directory = os.getcwd()
    path = os.path.join(directory,"6. CABANILLAS", "Mturn", "Muelle_optimo", "implantaciones.xlsx")
    regex = r'\b([^.]+)\.'

    df = pd.read_excel(path, sheet_name="ImplantaciónPicking", header=0)
    df = df.iloc[2:, 3:5]

    # Detectar y asignar los encabezados
    new_headers = df.iloc[0]  # La primera fila del DataFrame recortado
    df = df[1:]  # Eliminar la primera fila ahora que se usa como encabezado
    df.columns = new_headers  # Asignar los nuevos encabezados
    df['Ubi'] = df['Picking Location'].str.extract(regex)

    return df


    
def clasificador_LA_por_ubi(la):
    import pandas as pd
    import os
    
    directory = os.getcwd()
    
    path = os.path.join(directory,"6. CABANILLAS","Mturn","Muelle_optimo","implantaciones.xlsx")
    regex = r'^([^.]+)\.'

    df = pd.read_excel(path, sheet_name="ImplantaciónPicking", header=0)
    df = df.iloc[2:, 3:5]

    # Detectar y asignar los encabezados
    new_headers = df.iloc[0]  # La primera fila del DataFrame recortado
    df = df[1:]  # Eliminar la primera fila ahora que se usa como encabezado
    df.columns = new_headers  # Asignar los nuevos encabezados
    
    df['Ubi'] = df['Picking Location'].str.extract(regex)

    # Asegurarse de que 'LA' se trata como numérico para la comparación
    df['LA'] = pd.to_numeric(df['LA'], errors='coerce')

    # Filtrar el DataFrame para encontrar la fila donde 'LA' es igual al valor a buscar (la)
    filtro = df['LA'] == la
    if filtro.any():
        res = df.loc[filtro, 'Ubi'].values[0] #type:ignore
        return res
    else:
        return None
    
def consultar_vol(la):
    import pandas as pd
    import os
    directory = os.getcwd()
    columna = 8
    path = os.path.join(directory, "6. CABANILLAS", "Mturn", "Muelle_optimo", "Master_data.xlsx")
    
    df =pd.read_excel(path)
    # print("DataFrame cargado:")
    # print(df)
    
    filtro = df['Articulo'] == la
    # print("Valor de 'la':", la)
    # print("Valores de la columna 'Articulo':")
    # print(df['Articulo'])
    
    
    # Verificar si el valor 'la' se encuentra en la columna 'Articulo'
    if filtro.any():
        # Devolver la fila correspondiente al valor 'la'
        fila = df[filtro].iloc[0]
        valor_columna = fila[columna]
        return valor_columna
    else:
        # Si el valor 'la' no se encuentra, devolver None
        return None

def calcular_volumetria(la, units):
    # Aquí va tu lógica de cálculo de volumetría
    vol = consultar_vol(la) * units
    
    return vol

def procesar_dato_clave(data_item):
    key, value_list = data_item
    temp_list = []
    for value in value_list:
        la = value[0]
        units = value[1]

        # Obtener la ubicación usando la función clasificador_LA_por_ubi
        ubicacion = clasificador_LA_por_ubi(la)
        if ubicacion is not None:
            # Calcular la volumetría
            volumetria = calcular_volumetria(la, units)
            # Añadir la tupla (ubicación, volumetria) a la lista temporal
            temp_list.append((ubicacion, volumetria))

    return (key, temp_list)

def procesar_datos(data):
    import multiprocessing as mp
    # Crear un Pool de trabajadores
    pool = mp.Pool(mp.cpu_count())

    # Mapear los datos al Pool de trabajadores
    result_list = pool.map(procesar_dato_clave, data.items())

    # Cerrar el Pool y esperar a que terminen los trabajadores
    pool.close()
    pool.join()

    # Convertir la lista de resultados a un diccionario
    result_dict = dict(result_list)
    
    return result_dict

def ubicacion_con_mayor_volumen(result_dict):
    # Diccionario para sumar los volúmenes por ubicación
    suma_por_ubicacion = {}

    for key, ubicaciones in result_dict.items():
    
        for ubi, volum in ubicaciones:
            if ubi in suma_por_ubicacion:
                suma_por_ubicacion[ubi] += volum
            else:
                suma_por_ubicacion[ubi] = volum
        

    # Encontrar la ubicación con el mayor volumen
    max_ubi = max(suma_por_ubicacion, key=suma_por_ubicacion.get) #type: ignore
    return max_ubi, suma_por_ubicacion[max_ubi]

def obtener_r_mayor(diccionario):
    resultado = {}

    for key, value_list in diccionario.items():
        max_r = None
        max_volumetria = -float('inf')

        for ubicacion, volumetria in value_list:
            if volumetria > max_volumetria:
                max_r = ubicacion
                max_volumetria = volumetria

        resultado[key] = max_r

    return resultado

def selector_muelle_pasillo(location: str):
    if location  in ["R00", "R01"]:
        return 2
    elif location == "R01":
        return 3
    elif location == "R02":
        return 4
    elif location == "R03":
        return 5
    elif location == "R04":
        return 6
    elif location == "R05":
        return 7
    elif location == "R06":
        return 8
    elif location == "R07":
        return 9
    elif location == "R08":
        return 10
    elif location == "R09":
        return 11
    elif location in ["R10", "R11"]:
        return 14
    #elif location == "R11" or location[0] == "B":
    #    return 15
    elif str(location[0]) == "B":
        return 17
    else:
        return 11