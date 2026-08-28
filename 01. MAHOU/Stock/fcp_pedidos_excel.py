import os
import re

import base64
from alive_progress import alive_bar

import pandas as pd
import requests
#import openpyxl
from sqlalchemy import create_engine, text
from datetime import datetime

from pyfiglet import Figlet
from tabulate import tabulate
from colorama import Fore, Style, init

init(autoreset=True)

base_path = os.path.dirname(os.path.abspath(__file__))
almacen = 221

RAW_DIR = os.path.join(base_path, "raw_files")

def view_order(dataframe, pedido):

    if dataframe is None or dataframe.empty:
        print(Fore.RED + "\nNo hay información EDI para mostrar.")
        return

    pedido_buscado = str(pedido).strip()

    df_pedido = dataframe[
        dataframe["Pedido"].astype(str).str.strip() == pedido_buscado
    ].copy()

    if df_pedido.empty:
        print(Fore.RED + f"\nNo se encontró el pedido {pedido_buscado}.")
        return

    tabla = []

    for _, fila in df_pedido.iterrows():

        referencia = fila["Referencia"]

        if pd.isna(referencia) or str(referencia).strip() == "":
            referencia = "SIN REFERENCIA"

        tabla.append([
            fila["Línea"],
            referencia,
            fila["Tipo FCP"],
            fila.get("Valor FCP", "")
        ])

    figlet = Figlet(font="slant")

    print(Fore.CYAN)
    print(figlet.renderText("FCP PEDIDO"))

    print(
        Fore.WHITE
        + Style.BRIGHT
        + f"Pedido: {pedido_buscado}\n"
    )

    print(
        Fore.CYAN
        + tabulate(
            tabla,
            headers=["Línea", "Referencia", "Tipo FCP", "Valor"],
            tablefmt="fancy_grid",
            stralign="left",
            numalign="center"
        )
    )

    print(
        Fore.GREEN
        + Style.BRIGHT
        + f"\nTotal de líneas: {len(tabla)}"
    )

def obtener_escritorio():
    """Devuelve la ruta real del Escritorio de Windows, incluso si está redirigido a OneDrive."""
    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            CSIDL_DESKTOPDIRECTORY = 0x10
            SHGFP_TYPE_CURRENT = 0
            buffer = ctypes.create_unicode_buffer(wintypes.MAX_PATH)
            result = ctypes.windll.shell32.SHGetFolderPathW(
                None,
                CSIDL_DESKTOPDIRECTORY,
                None,
                SHGFP_TYPE_CURRENT,
                buffer,
            )
            if result == 0 and buffer.value:
                return buffer.value
        except Exception:
            pass

    # Fallback útil también si se prueba fuera de Windows
    return os.path.join(os.path.expanduser("~"), "Desktop")

def leer_excel(ruta_archivo):
    """Lee todos los albaranes de la columna A del Excel de entrada."""
    try:
        # header=None evita perder el primer albarán si el archivo no tiene cabecera.
        df = pd.read_excel(
            ruta_archivo,
            engine="openpyxl",
            sheet_name=0,
            usecols=[0],
            header=None,
            dtype=str,
        )

        df.columns = ["Albaran"]
        df["Albaran"] = df["Albaran"].astype(str).str.strip()
        df = df[
            df["Albaran"].notna()
            & (df["Albaran"] != "")
            & (df["Albaran"].str.lower() != "nan")
        ].copy()

        # Si la primera celda es una cabecera típica, la descartamos.
        cabeceras = {
            "albaran", "albarán", "albaranes",
            "pedido", "pedidos",
        }
        if not df.empty and df.iloc[0, 0].lower() in cabeceras: #type: ignore
            df = df.iloc[1:].copy()

        # Excel puede haber dejado un identificador numérico como 15885360.0.
        df["Albaran"] = df["Albaran"].str.replace(r"\.0$", "", regex=True)
        df = df.drop_duplicates().reset_index(drop=True)

        return df

    except FileNotFoundError:
        print(Fore.RED + f"❌ No se encuentra el archivo: {ruta_archivo}")
        return None
    except Exception as e:
        print(Fore.RED + f"❌ Error al leer {ruta_archivo}: {e}")
        return None

def guardar_excel_fcp(dataframe, ruta_salida):
    """Guarda exactamente los datos útiles de la consulta en un Excel formateado."""
    if dataframe is None or dataframe.empty:
        print(Fore.RED + "\n❌ No hay datos FCP para generar el Excel.")
        return False

    columnas = ["Pedido", "Línea", "Referencia", "Tipo FCP", "Valor FCP"]
    df_salida = dataframe.copy()

    for columna in columnas:
        if columna not in df_salida.columns:
            df_salida[columna] = ""

    df_salida = df_salida[columnas].copy()
    df_salida = df_salida.sort_values(
        by=["Pedido", "Línea"],
        kind="stable"
    ).reset_index(drop=True)

    try:
        with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
            df_salida.to_excel(writer, index=False, sheet_name="FCP Pedidos")

            ws = writer.sheets["FCP Pedidos"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # Formato sencillo y legible.
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.table import Table, TableStyleInfo

            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            borde = Side(style="thin", color="D9E2F3")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = Border(bottom=borde)

            # Tabla de Excel para filtros y bandas de filas.
            if ws.max_row >= 2:
                tabla = Table(displayName="TablaFCPPedidos", ref=ws.dimensions)
                estilo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                tabla.tableStyleInfo = estilo
                ws.add_table(tabla)

            anchos = {
                "A": 18,  # Pedido
                "B": 10,  # Línea
                "C": 22,  # Referencia
                "D": 24,  # Tipo FCP
                "E": 20,  # Valor FCP
            }
            for columna, ancho in anchos.items():
                ws.column_dimensions[columna].width = ancho

            ws.column_dimensions["B"].width = 10
            for cell in ws["B"]:
                cell.alignment = Alignment(horizontal="center")

        print(Fore.GREEN + Style.BRIGHT + f"\n✅ Excel generado: {ruta_salida}")
        print(Fore.GREEN + f"   Pedidos encontrados: {df_salida['Pedido'].nunique()}")
        print(Fore.GREEN + f"   Líneas procesadas: {len(df_salida)}")
        return True

    except PermissionError:
        print(Fore.RED + "\n❌ No se puede guardar el Excel. ¿Está abierto en Excel?")
        return False
    except Exception as e:
        print(Fore.RED + f"\n❌ Error generando el Excel: {type(e).__name__}: {e}")
        return False


def leer_sql_por_bloques(
    engine,
    valores,
    construir_query,
    params_fijos=None,
    tamano_bloque=2000,
):
    """
    Ejecuta una consulta SQL con una cláusula IN dividiendo los valores
    en bloques para no superar el límite de parámetros de SQL Server.

    Parámetros:
        engine:
            Motor SQLAlchemy.

        valores:
            Lista de valores que se utilizarán en el IN.

        construir_query:
            Función que recibe los placeholders y devuelve la consulta SQL.

        params_fijos:
            Parámetros adicionales como almacen, cliente, etc.

        tamano_bloque:
            Número máximo de elementos procesados en cada consulta.
    """

    params_fijos = params_fijos or {}

    # Limpiar valores nulos, vacíos y duplicados
    valores_limpios = []

    for valor in valores:
        if pd.isna(valor):
            continue

        valor_limpio = str(valor).strip()

        if not valor_limpio or valor_limpio.lower() == "nan":
            continue

        # Corrige posibles identificadores procedentes de Excel:
        # 23387339.0 -> 23387339
        valor_limpio = re.sub(r"\.0$", "", valor_limpio)

        valores_limpios.append(valor_limpio)

    # Elimina duplicados conservando el orden
    valores_limpios = list(dict.fromkeys(valores_limpios))

    if not valores_limpios:
        return pd.DataFrame()

    resultados = []
    total_bloques = (
        len(valores_limpios) + tamano_bloque - 1
    ) // tamano_bloque

    print(
        Fore.CYAN
        + f"Procesando {len(valores_limpios)} valores "
        + f"en {total_bloques} bloque(s)..."
    )

    with engine.connect() as conn:

        for numero_bloque, inicio in enumerate(
            range(0, len(valores_limpios), tamano_bloque),
            start=1,
        ):
            bloque = valores_limpios[
                inicio:inicio + tamano_bloque
            ]

            placeholders = ",".join(
                f":id{i}" for i in range(len(bloque))
            )

            params = {
                f"id{i}": valor
                for i, valor in enumerate(bloque)
            }

            # Añadir parámetros fijos como almacen
            params.update(params_fijos)

            query = text(construir_query(placeholders))

            print(
                Fore.WHITE
                + f"  Bloque {numero_bloque}/{total_bloques}: "
                + f"{len(bloque)} valores"
            )

            df_bloque = pd.read_sql(
                query,
                conn,
                params=params,
            )

            if not df_bloque.empty:
                resultados.append(df_bloque)

    if not resultados:
        return pd.DataFrame()

    resultado_final = pd.concat(
        resultados,
        ignore_index=True,
    )

    # Evita duplicados si un mismo identificador apareciese
    # en más de un bloque por cualquier motivo
    resultado_final = resultado_final.drop_duplicates(
        ignore_index=True
    )

    return resultado_final

def generar_nombres_unicos(filepath):
    base, ext = os.path.splitext(filepath)
    contador = 1

    while os.path.exists(filepath):
        filepath = f"{base}_{contador}{ext}"
        contador += 1
    return filepath

def generar_df_trabajo(df, almacen):

    engine = create_engine("mssql+pyodbc://@XGA_PROD")

    if df is None or df.empty:
        print(
            Fore.RED
            + "El archivo está vacío o no se pudo leer."
        )
        return []

    # Extraer albaranes únicos de la primera columna
    valores = (
        df.iloc[:, 0]
        .dropna()
        .astype(str)
        .str.strip()
        .loc[lambda serie: serie != ""]
        .unique()
        .tolist()
    )

    if not valores:
        print(Fore.RED + "No hay albaranes para consultar.")
        return []

    print(
        Fore.CYAN
        + Style.BRIGHT
        + f"\n1. Consultando {len(valores)} albaranes..."
    )

    # ---------------------------------------------------------
    # CONSULTA 1: OBTENER ID_DOC DE LOS ALBARANES
    # ---------------------------------------------------------

    def construir_query_iddoc(placeholders):
        return f"""
            SELECT
                AlbaranDoc,
                ID_Doc AS [ID_Doc]
            FROM
                vDocumentos
            WHERE
                ID_Cliente = 944
                AND ID_Almacen = :almacen
                AND CodigoTipoDocumento = 'ALB'
                AND AlbaranDoc IN ({placeholders})
        """

    df_2 = leer_sql_por_bloques(
        engine=engine,
        valores=valores,
        construir_query=construir_query_iddoc,
        params_fijos={"almacen": almacen},
        tamano_bloque=2000,
    )

    if df_2.empty:
        print(
            Fore.YELLOW
            + "No se encontraron documentos para los albaranes."
        )
        return []

    print(
        Fore.GREEN
        + f"   Documentos encontrados: {len(df_2)}"
    )

    # Extraer los ID_Doc
    ids_alb = (
        df_2["ID_Doc"]
        .dropna()
        .astype(str)
        .str.strip()
        .loc[lambda serie: serie != ""]
        .unique()
        .tolist()
    )

    if not ids_alb:
        print(
            Fore.YELLOW
            + "No se encontraron identificadores de documentos."
        )
        return []

    print(
        Fore.CYAN
        + Style.BRIGHT
        + f"\n2. Consultando subestados de "
        + f"{len(ids_alb)} documentos..."
    )

    # ---------------------------------------------------------
    # CONSULTA 2: OBTENER LOS SUBESTADOS DE LOS DOCUMENTOS
    # ---------------------------------------------------------

    def construir_query_subestados(placeholders):
        return f"""
            SELECT
                ID_Doc AS [ID_Doc],
                ID_SubEstadosDocumentos
                    AS [ID_SubEstadosDocumentos]
            FROM
                SubEstadosDocumentos
            WHERE
                ID_Cliente = 944
                AND ID_Doc IN ({placeholders})
        """

    df_3 = leer_sql_por_bloques(
        engine=engine,
        valores=ids_alb,
        construir_query=construir_query_subestados,
        tamano_bloque=2000,
    )

    if df_3.empty:
        print(
            Fore.YELLOW
            + "No se encontraron subestados para los documentos."
        )
        return []

    print(
        Fore.GREEN
        + f"   Subestados encontrados: {len(df_3)}"
    )

    # Unir documentos con subestados
    df_23 = df_2.merge(
        df_3,
        on="ID_Doc",
        how="inner",
    )

    if df_23.empty:
        print(
            Fore.YELLOW
            + "No se pudieron relacionar documentos y subestados."
        )
        return []

    # Extraer ID_SubEstadosDocumentos
    id_subest = (
        df_23["ID_SubEstadosDocumentos"]
        .dropna()
        .astype(str)
        .str.strip()
        .loc[lambda serie: serie != ""]
        .unique()
        .tolist()
    )

    if not id_subest:
        print(
            Fore.YELLOW
            + "No hay identificadores de subestados para consultar."
        )
        return []

    print(
        Fore.CYAN
        + Style.BRIGHT
        + f"\n3. Consultando ficheros de "
        + f"{len(id_subest)} subestados..."
    )

    # ---------------------------------------------------------
    # CONSULTA 3: OBTENER LAS RUTAS DE LOS FICHEROS
    # ---------------------------------------------------------

    def construir_query_ficheros(placeholders):
        return f"""
            SELECT
                ID_SubEstadosDocumentos
                    AS [ID_SubEstadosDocumentos],
                NombreFicheroBackupSubEstadoTransmision
            FROM
                SubEstadosTransmision
            WHERE
                ID_SubEstadosDocumentos
                    IN ({placeholders})
        """

    df_4 = leer_sql_por_bloques(
        engine=engine,
        valores=id_subest,
        construir_query=construir_query_ficheros,
        tamano_bloque=2000,
    )

    if df_4.empty:
        print(
            Fore.YELLOW
            + "No se encontraron ficheros asociados "
            + "a los subestados."
        )
        return []

    print(
        Fore.GREEN
        + f"   Ficheros encontrados: {len(df_4)}"
    )

    # Unir la información completa
    df_234 = df_23.merge(
        df_4,
        on="ID_SubEstadosDocumentos",
        how="inner",
    )

    if df_234.empty:
        print(
            Fore.YELLOW
            + "No se pudieron relacionar los subestados "
            + "con los ficheros."
        )
        return []

    rutas = (
        df_234["NombreFicheroBackupSubEstadoTransmision"]
        .dropna()
        .astype(str)
        .str.strip()
        .loc[lambda serie: serie != ""]
        .unique()
        .tolist()
    )

    print(
        Fore.GREEN
        + Style.BRIGHT
        + f"\n✅ Rutas de ficheros obtenidas: {len(rutas)}"
    )

    return rutas

def extraer_nombre_fichero(ruta, content_disposition=''):

    # 1. Intentar desde header
    if "filename=" in content_disposition:
        raw_filename = content_disposition.split("filename=")[-1].strip().strip('"')
        nombre = os.path.basename(raw_filename)

    else:
        # 2. Sacar desde ruta
        nombre = os.path.basename(ruta.replace("\\", "/"))

    # 3. Validar que empiece por I1084
    if not nombre.startswith("I1084"):
        return None  # 👈 clave para filtrar

    # 4. Limpiar caracteres problemáticos
    nombre = nombre.replace("\\", "_")

    return nombre

def descargar_edi(df, almacen):
    limpiar_raw_files(RAW_DIR)
    user = "JGMERAS"
    password = "M1j3kMICrdmxlRFVY0g1"

    base_url = "http://10.19.16.125"
    download_path = "/fga/MtoDocumentosTr/DescargaFicheroSubestado"

    if not os.path.exists(RAW_DIR):
        print(f"Creando carpeta: {RAW_DIR}")
        os.makedirs(RAW_DIR, exist_ok=True)

    ruta_list = generar_df_trabajo(df, almacen)

    # Si no hay rutas para procesar
    if ruta_list is None:
        print("No hay rutas para procesar.")
        return pd.DataFrame()

    if isinstance(ruta_list, pd.DataFrame):
        if ruta_list.empty:
            print("No hay rutas para procesar.")
            return pd.DataFrame()
    else:
        if len(ruta_list) == 0:
            print("No hay rutas para procesar.")
            return pd.DataFrame()
    
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0"
    })

    try:
        #hacemos login con la primera opción
        primera_ruta = ruta_list[0]
        primera_url = f"{base_url}{download_path}?nombrefichero={primera_ruta}"

        r1 = session.get(primera_url, allow_redirects=True, timeout=30)

        login_url = r1.url
        payload = {
            "Nombre" : user,
            "password": password,
            "ActualizarPassword": "False",
            "Agenda": "False",
            "Error": "",
            "Login": "Aceptar",
        }
        headers = {
            "Referer": r1.url,
            "Content-Type": "application/x-www-form-urlencoded",
            "User-Agent": "Mozilla/5.0"
        }

        r2 = session.post(login_url, data=payload, headers=headers, allow_redirects=True, timeout=30)

        for ruta in ruta_list:
            try:
                ruta_codificada = (ruta) #type:ignore
                url = f"{base_url}{download_path}?nombrefichero={ruta_codificada}"

                r3 = session.get(url, stream=True, timeout=60)
                #print("Descarga: ",r3.status_code, url)
                content_type = (r3.headers.get("Content-Type") or "").lower()
                content_disposition = r3.headers.get("Content-Disposition", "")

                if r3.status_code == 200:
                    filename = extraer_nombre_fichero(ruta, content_disposition)
                    # Si la ruta no corresponde a un I1084, la ignoramos
                    if not filename:
                        continue

                    filepath = os.path.join(RAW_DIR, filename)
                    filepath = generar_nombres_unicos(filepath)

                    with open(filepath, "wb") as f:
                        for chunk in r3.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    #print("Archivo guardado")
                else:
                    with open("error_response.html", "w", encoding="utf-8") as f:
                            f.write(r3.text)

                            #print("Descarga erronea, se obtuvo HTML en vez del archivo")
            except Exception as e:
                print(Fore.YELLOW + f"⚠️ Error descargando {ruta}: {type(e).__name__}: {e}")
                continue
                

    finally:
        session.close()

    # Una vez descargados los EDI, los procesamos y devolvemos el DataFrame
    return process_files(RAW_DIR)

def process_files(input_folder):
    filas = []

    files = [
        f for f in os.listdir(input_folder)
        if os.path.isfile(os.path.join(input_folder, f))
    ]

    total_files = len(files)

    with alive_bar(total_files, title="Procesando archivos") as bar:

        for filename in files:

            filepath = os.path.join(input_folder, filename)

            edi_text = decode_file(filepath)

            if edi_text is None:
                bar()
                continue

            data = extract_segments(edi_text)
            pedido = data["pedido"]

            for numero_linea, linea in enumerate(data["lineas"], start=1):

                filas.append({
                    "Pedido": pedido,
                    "Línea": numero_linea,
                    "Referencia": linea["referencia"],
                    "Tipo FCP": linea["tipo_fcp"],
                    "Valor FCP": linea["valor_fcp"],
                    "Archivo": filename
                })

            bar()

    return pd.DataFrame(filas)

def extract_base64(lines):
    """
    Encuentra dónde empieza el contenido base64.
    Primero busca 'Content-Disposition:'.
    Si no lo encuentra, empieza desde la línea 5.
    """
    start_index = None

    for i, line in enumerate(lines):
        if "Content-Disposition:" in line:
            start_index = i + 1
            break

    if start_index is None:
        start_index = 5

    base64_data = "".join(lines[start_index:]).strip()
    return base64_data

def decode_file(filepath):
    
    with open(filepath, "rb") as f:
        raw = f.read()

    # 1) intentar leer como texto directamente
    text = raw.decode("utf-8", errors="ignore")

    # Si ya parece EDI, no decodificamos
    if "UNB" in text or "UNH" in text or "LIN+" in text:
        return text

    # 2) si no, intentar decodificar base64
    lines = text.splitlines()

    candidate_lines = []
    in_body = False

    for line in lines:
        line_strip = line.strip()

        #Saltar cabeceras MIME hasta la primera línea vacía
        if not in_body:
            if line_strip == "":
                in_body = True
            continue

        # Cortamos si llega boundary final
        if line_strip.startswith("--"):
            break

        #Nos quedamos con base64
        if re.fullmatch(r"[A-Za-z0-9+/=]+", line_strip):
            candidate_lines.append(line_strip)
    
    base64_data = "".join(candidate_lines)

    if not base64_data:
        print("⚠️ No se encontró contenido base64 válido.")
        print(f"Ruta: {filepath}")
        print("Primeras 200 chars:", repr(text[:200]))
        return None

    try:
        decoded_bytes = base64.b64decode(base64_data, validate=True)
    except Exception as e:
        print("❌ Error decodificando base64:")
        print(f"Ruta: {filepath}")
        print(f"Error: {e}")
        print("Base64 data (primeros 200 chars):", repr(base64_data[:200]))
        return None
    
    edit_text = decoded_bytes.decode("utf-8", errors="ignore")

    if "LIN+" not in edit_text:
        print("⚠️ El texto decodificado no parece ser un archivo EDI válido.")
        print(f"Ruta: {filepath}")
        print("Primeras 200 chars:", repr(edit_text[:200]))
        return None
    
    return edit_text
    
def extract_segments(edi_text):
    result = {
        "pedido": None,
        "lineas": []
    }

    # Pedido (cabecera). Admite BGM+80E:...+PEDIDO y BGM+80E+PEDIDO
    match = re.search(r"BGM\+80E(?::[^+]*)?\+([^+']+)", edi_text)
    if match:
        result["pedido"] = match.group(1).strip()

    # Cada bloque LIN corresponde a una referencia/línea del pedido
    bloques = re.split(r"(?=LIN\+)", edi_text)

    for bloque in bloques:
        if not bloque.startswith("LIN+"):
            continue

        referencia = None
        tipo_fcp = "SIN FCP"
        valor_fcp = ""

        pia = re.search(r"PIA\+1\+([^:+']+):SA", bloque)
        if pia:
            referencia = pia.group(1).strip()

        # Mantenemos la misma lógica del script original, pero mostrando
        # también el valor asociado para que la consulta sea más útil.
        dtm264 = re.search(r"DTM\+264:(\d{8})", bloque)
        dtm267 = re.search(r"DTM\+267:(\d{8})", bloque)
        rff_fcp = re.search(r"RFF\+FCP:([^']+)", bloque)

        if dtm264:
            tipo_fcp = "FCP DESDE"
            valor_fcp = datetime.strptime(dtm264.group(1), "%Y%m%d").strftime("%d/%m/%Y")
        elif dtm267:
            tipo_fcp = "FCP EXACTA"
            valor_fcp = datetime.strptime(dtm267.group(1), "%Y%m%d").strftime("%d/%m/%Y")
        elif rff_fcp:
            tipo_fcp = "FCP POR PORCENTAJE"
            valor_fcp = rff_fcp.group(1).strip()

        result["lineas"].append({
            "referencia": referencia,
            "tipo_fcp": tipo_fcp,
            "valor_fcp": valor_fcp
        })

    return result

def limpiar_raw_files(folder_path):
    if not os.path.exists(folder_path):
        print(f"La carpeta no existe: {folder_path}")
        return

    archivos_eliminados = 0

    for filename in os.listdir(folder_path):
        filepath = os.path.join(folder_path, filename)

        if os.path.isfile(filepath):
            try:
                os.remove(filepath)
                archivos_eliminados += 1
            except Exception as e:
                print(f"Error eliminando {filename}: {e}")

    print(f"Archivos eliminados: {archivos_eliminados}")

def procesar_archivo_muestra():
    escritorio = obtener_escritorio()
    ruta_entrada = os.path.join(escritorio, "Archivo_muestra.xlsx")

    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta_salida = os.path.join(escritorio, f"FCP_Pedidos{fecha}.xlsx")

    print(Fore.CYAN + Style.BRIGHT + "\n=== FCP PEDIDOS ===")
    print(f"Archivo de entrada: {ruta_entrada}")

    df_albaranes = leer_excel(ruta_entrada)

    if df_albaranes is None or df_albaranes.empty:
        print(Fore.RED + "\n❌ No hay albaranes para procesar en la columna A.")
        return

    print(Fore.WHITE + Style.BRIGHT + f"Albaranes leídos: {len(df_albaranes)}")

    resultados = descargar_edi(df_albaranes, almacen)

    if resultados is None or resultados.empty:
        print(Fore.RED + "\n❌ No se ha podido obtener información FCP de los albaranes.")
        return

    guardar_excel_fcp(resultados, ruta_salida)


if __name__ == "__main__":
    procesar_archivo_muestra()
